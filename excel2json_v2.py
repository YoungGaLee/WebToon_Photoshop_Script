'''
excel2json.py

엑셀파일 내의 대사를 읽어오고, 레이어 생성할 위치를 계산하여 json형태로 저장시키는 코드

마지막 수정일 : 2021-12-06
이가영
'''

from psd_tools import PSDImage
from openpyxl import load_workbook
import numpy as np
import json
import natsort
import os
from natsort import os_sorted

def addressPSD():
    """
    폴더 내의 파일들 중 확장자가 psd, psb인 파일에 대한 절대주소, 이름 가져오는 함수

    .. note:: json파일에 저장되는 정보
              ExtendScript가 익숙치 않아서 py파일에서 작업함.
             절대주소 - script에서 psd파일을 열기 위해서
             이름 - json데이터의 key. 이를 통해 script에서 value에 접근함.

    :return: `str` 엑셀파일절대주소(excelPath), `list` psd(b)파일의 절대주소(psdPath), `list` psd(b)파일이름(file_name)
    """
    path_split = os.getcwd().split(os.path.sep)
    del path_split[0]

    # ExtendScript에서는 window임에도 불구하고 폴더구분자가 '/'이다.
    path_jsx = '/'+'/'.join(path_split)

    path_dir = './'
    file_list = os_sorted(os.listdir(path_dir))

    psdPath=[]
    file_name = []
    excelPath = ''
    for path in (file_list):
        if path.startswith('~$'):
            continue

        if os.path.splitext(path)[1] == '.xlsx':
            assert excelPath == '', '폴더 내 엑셀파일이 두개 이상입니다. 확인해주세요.'
            excelPath = path_jsx + '/' + path

        if os.path.splitext(path)[1] == '.psd' or os.path.splitext(path)[1] =='.psb':
            file_path = path_jsx + '/' + path
            file_name.append(path)
            psdPath.append(file_path)

    sorted_file_name = natsort.natsorted(file_name)
    return excelPath, psdPath, sorted_file_name


def heigthPSD(pathList):
    """
    ratio 계산을 위해 모든 psd파일의 height를 합친 전체 height값과 각 psd파일의 height에 대한 list를 구하는 함수

    .. note:: 파일이름(pathList)를 통해 psd파일을 열고 그때 height에 대해서 계산한다. (psd_tools API 사용)

    :param pathList: `list` psd(b)파일이름
    :return: `int` 작품 전체 psd의 height 값(height), `list` 각 psd(b)의 height(height_list)
    """
    global webtoon_width
    height = 0
    height_list = []
    for path in pathList:
        psd = PSDImage.open(path)
        height_list.append(psd.height)
        height += psd.height
    webtoon_width = psd.width
    return height, height_list


def heightExcel():
    """
    ratio 계산을 위해 전체 엑셀이미지가 몇번째 줄에서 끝나는지 구하는 함수

    .. note:: 엑셀이미지의 마지막 이미지는 psd파일에 포함되지 않음. 따라서 ratio 계산을 위해서는 똑같이 엑셀의 마지막 이미지를 포함하지 않은 상태로 row를 구해야함.
              한마지로 엑셀이미지의 [-2]번째 이미지가 몇번째 row에서 끝나는지 그 값을 구하는 코드
              Object.anchor._from.row는 해당 이미지가 위치한 첫번째 row를 반환하는 메소드로 [-1]번째 이미지의 반환값을 통해 [-2]번째 이미지의 마지막 row값을 구함.

    :return: `int` 마지막 row 값(rowoff)
    """
    web_imgs_excel = excel_sheet._images
    rowoff = web_imgs_excel[-1].anchor._from.row
    final = web_imgs_excel[-1].height * (web_imgs_excel[1].anchor._from.row / web_imgs_excel[0].height)
    webtoon_rowoff = rowoff + final
    return round(webtoon_rowoff)


if __name__ == "__main__":
    excel_path, psd_path, file_name = addressPSD()           # 파일 path 정리
    webtoon_height, psd_height_list = heigthPSD(file_name)   # psd파일의 길이 계산
    print('총 PSD height : ',webtoon_height)

    # 엑셀 읽어오기
    excel = load_workbook(excel_path, data_only=True)
    excel_sheet = excel.active
    webtoon_excel_height = heightExcel()      # 엑셀파일의 -2번째 이미지까지의 길이
    print('엑셀 마지막 row :',webtoon_excel_height)

    # 엑셀내 대사들이 위치한 col, row 범위
    max_col = excel_sheet.max_column
    min_col = excel_sheet.min_column
    max_row = excel_sheet.max_row

    # ratio 계산
    """
    엑셀내의 전체 이미지와 대사들에 대해서 각 psd에 해당하는 범위를 나눠주기 위해서 ratio를 계산한다.
    각 psd파일의 height를 엑셀의 row 개수로 변환하여 해당 row범위에 해당하는 대사를 json파일에 넣어준다.
    ex ) 01.psd의 height가 12311이라면 이는 엑셀내의 286개의 row에 해당하고 해당 row 범위에 있는 대사들을 json파일에 '01.psd'라는 key값에 대한 value로 저장한다.  
    """
    psd_row_list = np.around(np.array(psd_height_list) * (webtoon_excel_height / webtoon_height))
    every_ratio_row2psd = np.array(psd_height_list) / np.array(psd_row_list)

    # JSON파일 생성하기
    with open('jsonFile.json', 'w', encoding='UTF-8') as w:
        json_file = {'psdPath': psd_path, 'psdName': file_name}

        # psd파일 이름으로 key 생성
        for psd_name in file_name:
            json_file[psd_name] = []

        count = 0
        json_psd = []

        # 엑셀 읽어오고 json에 저장해주기
        '''
        엑셀을 한줄한줄 읽어오는 과정에서 각 psd 범위(row로 표현된)를 참고하여 json에 저장한다.
        psd_row_list : 각 psd가 엑셀내에서 몇개의 줄에 해당하는지 나타낸 리스트
        ex) 처음부터 끝까지 내용을 읽어오는데 286번 row까지는 01.psd에 대한 대사에 해당하고, 그 이후 546개 줄은 02.psd에 대한 대사에 해당한다. 
        '''
        row_list = []
        last_context_count = 0
        exit_col = 0
        condition = 0   # 윤문여부

        # 윤문이 있는 작품인지 확인해보기 (condition)
        for i in range(min_col, max_col + 1):
            context = excel_sheet.cell(row=1, column=i).value
            if context == '윤문':
                condition = 1

        '''
        윤문이 있으면, condition = 1 이며 윤문의 context들을 대사로 읽어냄.
        윤문이 없다면, condition = 0 이며 '번역'의 context들을 대사로 읽어냄.
        '번역' 등 존재하지 않을 경우, 제외할 context 없이 대사로 읽어냄.
        '''

        # 값 읽어오기
        print(min_col,max_col)

        for c in range(min_col, max_col + 1):

            boundary = 0 # 비고 등 다음줄로 가면 boundary가 0으로 끝나지
            context = excel_sheet.cell(row=1, column=c).value

            # context 중 제외할 부분 (break)
            if context == '비고': break     # 비고 값은 제외합니다.
            if context == '번역':
                if condition == 1: break  # 윤문이 있을 경우, 번역 값은 제외합니다.
                else: pass
            if context == '윤문': pass

            for r in range(1, max_row + 1):
                context = excel_sheet.cell(row=r, column=c).value
                if boundary == len(file_name):
                    continue

                # psd 범위 나누기 (다음 psd로 넘어가기)
                if count >= psd_row_list[boundary]:
                    # 새로 해당 psd 리스트 생성 (json)
                    if json_file[file_name[boundary]] == []:
                        json_file[file_name[boundary]] = json_psd
                        row_list.append(r)
                    # 읽어야 할 col이 둘 이상일 경우 (이미 존재하는 psd리스트에 context 추가)
                    else: json_file[file_name[boundary]].extend(json_psd)

                    count = 0
                    boundary += 1
                    json_psd = []

                count += 1
                if context == None or context == '윤문' or context == '번역':  # 레이어 생성에 있어 제외
                    continue

                # ExtendScript에서는 줄바꿈 기호를 '\n'이 아닌 '\r'을 사용한다.
                try :
                    if '\n' in context:
                        context = context.replace('\n', '\r')
                except TypeError:  # 문자가 아닌 숫자로 읽는 경우
                    context = str(context)
                    context = context.replace('\n', '\r')

                # 레이어 생성할 위치값
                psd_x = webtoon_width // 4
                psd_y = count * every_ratio_row2psd[boundary]

                # v2 추가된 부분
                # v2 : 엑셀의 다음 row에 적는 방식으로 줄바꿈하는 경우
                if (count - 1) == last_context_count:    
                    try:
                        text = json_psd[-1]['text']
                        ans = text + '\r' + context
                        json_psd[-1]['text'] = ans
                    except IndexError: # 맨처음
                        last_context_count = 0
                        line = [['text', context], ['x', round(psd_x, 2)], ['y', round(psd_y, 2)]]
                        json_psd.append(dict(line))

                else:
                    last_context_count = 0
                    line = [['text', context], ['x', round(psd_x, 2)], ['y', round(psd_y, 2)]]
                    json_psd.append(dict(line))

                last_context_count = count
                psd_x += webtoon_width // 2

            # 마지막 row로 인한 IndexError처리 (boundary range)
            if boundary == len(file_name):
                if context == None: pass
                else: line = [['text', context], ['x', round(psd_x, 2)], ['y', round(psd_y, 2)]]
            else: json_file[file_name[boundary]] = json_psd

        row_list.insert(0, 1)
        print('\n엑셀에서의 각 PSD의 시작 row 값 : ', row_list)

        json.dump(json_file, w, indent=4, ensure_ascii=False)
        print("json파일 생성완료")
