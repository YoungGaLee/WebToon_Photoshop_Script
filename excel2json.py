'''
excel2json.py

엑셀파일 내의 대사를 읽어오고, 레이어 생성할 위치를 계산하여 json형태로 저장시키는 코드

마지막 수정일 : 2021-11-06
이가영
'''
import os
from psd_tools import PSDImage
from openpyxl import load_workbook
import numpy as np
import json
import glob
import natsort    # pip install natsort

def addressPSD():
    """
    폴더 내의 파일들 중 확장자가 psd, psb인 파일에 대한 절대주소, 이름 가져오는 함수

    .. note:: json파일에 저장되는 정보
              ExtendScript가 익숙치 않아서 py파일에서 작업함.
             절대주소 - script에서 psd파일을 열기 위해서
             이름 - json데이터의 key. 이를 통해 script에서 value에 접근함.

    :return: `str` 엑셀파일절대주소(excelPath), `list` psd(b)파일의 절대주소(psdPath), `list` psd(b)파일이름(file_name)
    """
    path_split = os.getcwd().split(os.path.sep)
    del path_split[0]

    # ExtendScript에서는 window임에도 불구하고 폴더구분자가 '/'이다.
    path_jsx = '/'+'/'.join(path_split)

    path_dir = './'
    file_list = os.listdir(path_dir)


    # print(file_list,"sort어떻게 되고 있는거지")

    # fl = sorted(glob.glob('*'), key=os.path.getmtime)
    fl = sorted(glob.glob('*'),reverse=True)
    # print(fl, "둘이 방식이 다른가")

    psdPath=[]
    file_name = []
    print(file_list)
    excelPath = ''
    for path in (file_list):
        if path.startswith('~$'):
            continue

        if os.path.splitext(path)[1] == '.xlsx':
            assert excelPath == '', '폴더 내 엑셀파일이 두개 이상입니다. 확인해주세요.'
            excelPath = path_jsx + '/' + path


        if os.path.splitext(path)[1] == '.psd' or os.path.splitext(path)[1] =='.psb':
            file_path = path_jsx + '/' + path
            file_name.append(path)
            psdPath.append(file_path)

    sorted_file_name = natsort.natsorted(file_name)

    return excelPath, psdPath, sorted_file_name




def heigthPSD(pathList):
    """
    ratio 계산을 위해 모든 psd파일의 height를 합친 전체 height값과 각 psd파일의 height에 대한 list를 구하는 함수

    .. note:: 파일이름(pathList)를 통해 psd파일을 열고 그때 height에 대해서 계산한다. (psd_tools API 사용)

    :param pathList: `list` psd(b)파일이름
    :return: `int` 작품 전체 psd의 height 값(height), `list` 각 psd(b)의 height(height_list)
    """
    global webtoon_width
    height = 0
    height_list = []
    for path in pathList:
        psd = PSDImage.open(path)
        height_list.append(psd.height)
        height += psd.height
    webtoon_width = psd.width

    return height, height_list


def heightExcel():
    """
    ratio 계산을 위해 전체 엑셀이미지가 몇번째 줄에서 끝나는지 구하는 함수

    .. note:: 엑셀이미지의 마지막 이미지는 psd파일에 포함되지 않음. 따라서 ratio 계산을 위해서는 똑같이 엑셀의 마지막 이미지를 포함하지 않은 상태로 row를 구해야함.
              한마지로 엑셀이미지의 [-2]번째 이미지가 몇번째 row에서 끝나는지 그 값을 구하는 코드
              Object.anchor._from.row는 해당 이미지가 위치한 첫번째 row를 반환하는 메소드로 [-1]번째 이미지의 반환값을 통해 [-2]번째 이미지의 마지막 row값을 구함.

    :return: `int` 마지막 row 값(rowoff)
    """
    web_imgs_excel = excel_sheet._images
    print()
    rowoff = web_imgs_excel[-1].anchor._from.row
    final = web_imgs_excel[-1].height * (web_imgs_excel[1].anchor._from.row / web_imgs_excel[0].height)
    webtoon_rowoff = rowoff + final
    # print(rowoff,"rowoff")
    # print(webtoon_rowoff,"webtoon_rowoff")

    return round(webtoon_rowoff)


if __name__ == "__main__":
    excel_path, psd_path, file_name = addressPSD()         # 파일 path 정리
    webtoon_height, psd_height_list = heigthPSD(file_name) # psd파일의 길이 계산
    print(file_name)
    print(excel_path)
    # print(psd_height_list)
    # print(webtoon_height)



    # 엑셀 읽어오기
    excel = load_workbook(excel_path, data_only=True)
    excel_sheet = excel.active
    webtoon_excel_height = heightExcel()      # 엑셀파일의 -2번째 이미지까지의 길이

    # 엑셀내 대사들이 위치한 col, row 범위
    max_col = excel_sheet.max_column
    min_col = excel_sheet.min_column
    max_row = excel_sheet.max_row

    # ratio 계산
    """
    엑셀내의 전체 이미지와 대사들에 대해서 각 psd에 해당하는 범위를 나눠주기 위해서 ratio를 계산한다.
    각 psd파일의 height를 엑셀의 row 개수로 변환하여 해당 row범위에 해당하는 대사를 json파일에 넣어준다.
    ex ) 01.psd의 height가 12311이라면 이는 엑셀내의 286개의 row에 해당하고 해당 row 범위에 있는 대사들을 json파일에 '01.psd'라는 key값에 대한 value로 저장한다.  
    """
    psd_row_list = np.around(np.array(psd_height_list) * (webtoon_excel_height / webtoon_height))
    every_ratio_row2psd = np.array(psd_height_list) / np.array(psd_row_list)

    # print(webtoon_excel_height,"webtoon_excel_height")
    # print(webtoon_height,"webtoon_height")
    # print(psd_height_list,"psd_height_list")
    # print(psd_row_list,"psd_row_list")
    # print(every_ratio_row2psd,"every_ratio_row2psd")


    # JSON파일 생성하기
    with open('jsonFile.json', 'w', encoding='UTF-8') as w:
        json_file = {'psdPath': psd_path, 'psdName': file_name}

        # psd파일 이름으로 key 생성
        for psd_name in file_name:
            json_file[psd_name] = []

        count = 0
        boundary = 0
        json_psd = []

        # 엑셀 읽어오고 json에 저장해주기
        '''
        엑셀을 한줄한줄 읽어오는 과정에서 각 psd 범위(row로 표현된)를 참고하여 json에 저장한다.
        psd_row_list : 각 psd가 엑셀내에서 몇개의 줄에 해당하는지 나타낸 리스트
        ex) 처음부터 끝까지 내용을 읽어오는데 286번 row까지는 01.psd에 대한 대사에 해당하고, 그 이후 546개 줄은 02.psd에 대한 대사에 해당한다. 
        '''
        row_list = []
        # print(max_row,"max_row")
        for r in range(1, max_row + 1):


            # print(r,max_row)

            print(psd_row_list,"psd_row_list")

            if count >= psd_row_list[boundary]:
                # print(psd_row_list,psd_row_list[boundary])
                json_file[file_name[boundary]] = json_psd
                row_list.append(r)

                count = 0
                boundary += 1
                json_psd = []

            psd_x = webtoon_width // 4
            count += 1

            for c in range(min_col, max_col + 1):
                context = excel_sheet.cell(row=r, column=c).value

                print(r, file_name[boundary], context) # r이라는 열에 context
                print(row_list, "row_list")
                # print(r, file_name[boundary],"얍")
                # print(boundary,context)
                if context == None:
                    continue


                # ExtendScript에서는 줄바꿈 기호를 '\n'이 아닌 '\r'을 사용한다.
                try :
                    if '\n' in context:
                        context = context.replace('\n', '\r')
                except TypeError:
                    context = str(context)
                    context = context.replace('\n', '\r')
                    # 20:37:00

                # print(psd_row_list,psd_row_list[boundary])
                # print(every_ratio_row2psd,every_ratio_row2psd[boundary-1])

                psd_y = count * every_ratio_row2psd[boundary]

                line = [['text', context], ['x', round(psd_x, 2)], ['y', round(psd_y, 2)]]
                json_psd.append(dict(line))
                psd_x += webtoon_width // 2

        json_file[file_name[boundary]] = json_psd

        json.dump(json_file, w, indent=4, ensure_ascii=False)

        print("json파일 생성완료")

